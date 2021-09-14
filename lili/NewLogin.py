import os
import random
import threading
import traceback
import datetime
from concurrent.futures import ThreadPoolExecutor

from selenium import webdriver
import time
import encodings.idna

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.select import Select
import openpyxl

#用于个人登陆
class login:

    lock=threading.Lock()

    s1 = threading.Semaphore(10)

    # 均分数组
    def list_split(self, n):
        return [self[i:i + n] for i in range(0, len(self), n)]

    # 登录页面
    def loginWebOne(self, ac,school,professional,mynumber):

        # self.s1.acquire()

        ac = ('15288846412','haha1688')


        print("编号: %s -> 账号: %s 登陆开始======================》" % (mynumber,ac[0]))


        # 笔记本路径
        root_path = os.path.abspath(os.path.dirname(__file__)).split('lili')[0]
        chromedriver = root_path+'images\chromedriver.exe'

        # json = requests.get("http://127.0.0.1:5000/get/").json()['proxy']

        myoptions = Options()

        myoptions.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"')
        # 设置代理

        # myoptions.add_argument('--proxy-server=' + json)

        myoptions.add_argument("--incognito")

        # myoptions.add_argument('--headless')

        myoptions.add_experimental_option("excludeSwitches", ['enable-automation', 'enable-logging'])

        # myoptions.add_argument("--proxy-server=http://27.191.60.197:3256")

        url = "http://jinanzyk.36ve.com/login/login"



        # 打开一个浏览器窗口
        driver = webdriver.Chrome(
            executable_path=chromedriver,
            options=myoptions)
        driver.implicitly_wait(20)
        driver.set_window_size(800, 800)
        # driver.maximize_window()

        try:
            # 发送请求
            driver.get(url)
            time.sleep(4)

            # 登录
            driver.find_element_by_id("loginform-username").send_keys(ac[0])
            time.sleep(2)

            driver.find_element_by_id("loginform-password").send_keys(ac[1])
            time.sleep(2)

            driver.find_element_by_xpath('//*[@id="login-form-1"]/div[3]/button').click()

            nowurl = driver.current_url

            # 登录成功与否判断
            if nowurl == url:
                print("账号: %s 登陆失败" % ac[0])
                result_login=login.isElementPresent(self,driver,'xpath','//*[@id="login-form-1"]/div[1]/div[2]/div[2]/p')

                if result_login:
                    login.writeExcel(self,ac[0])
                    time.sleep(10)

            else:
                print("账号: %s 登陆成功" % ac[0])
                time.sleep(2)

                # edit=login.isElementPresent(self,driver,'xpath','/html/body/div[4]/div/div/div/div[2]/a')

                # 判断是否需要个人完善信息
                # if edit:
                #     login.PerfectInformation(self,driver,ac,school,professional,mynumber)

                # time.sleep(1)
                print("账号: %s 浏览资源***************************" % ac[0])
                login.BrowseTheResources(self, driver)


        except Exception as e:
            print("======================账号: %s 账号异常信息===============================" % ac[0])
            print(e)
            pass
        finally:
            time.sleep(random.randint(60,70))
            # time.sleep(random.randint(20,40))
            print("~~~~~~~~~~~~~账号: %s 登录结束，关闭页面~~~~~~~~~~~~~~~~~" % ac[0])
            driver.delete_all_cookies()
            driver.quit()
            # self.s1.release()




    """
    完善个人信息
    """
    def PerfectInformation(self,driver,ac,school,professional,mynumber):

        driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[2]/a').click()

        time.sleep(3)
        driver.find_element_by_id("edit_info").click()
        time.sleep(2)

        # 设置名字
        name = driver.find_element_by_id("userinfo-pet_name").get_attribute('value')
        print("获取的名字为：%s" % name)
        if name == '' or name is None:
            driver.find_element_by_id("userinfo-pet_name").send_keys(
                random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
        time.sleep(2)

        # 设置生日
        month = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"]
        day = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"
            , "11", "12", "13", "14", "15", "16", "17", "18", "19", "20"
            , "21", "22", "23", "24", "25", "26", "27", "28", "29", "30"]
        driver.find_element_by_id("userinfo-birthday").send_keys(
            "2001-%s-%s" % (random.choice(month), random.choice(day)))
        time.sleep(2)

        # 区分身份
        name_flag = driver.find_element_by_xpath('//*[@id="userinfo-type"]').get_attribute('value')
        print("%s 获取的身份为：%s" % (ac[0], name_flag))

        if name_flag == "社会学习者":
            driver.find_element_by_id("userinfo-company").send_keys(
                random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
            time.sleep(2)
            driver.find_element_by_id("userinfo-pet_name").send_keys(
                random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
            time.sleep(2)
            # 设置详细地址
            driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
            time.sleep(2)

            driver.find_element_by_xpath('//*[@id="form-1"]/div[14]/button').click()

        if name_flag == "企业用户":
            driver.find_element_by_id("userinfo-company").send_keys(
                random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
            time.sleep(2)
            # 设置详细地址
            driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
            time.sleep(2)

            driver.find_element_by_xpath('//*[@id="form-1"]/div[14]/button').click()

        if name_flag == "学生":
            # 设置入校年份
            Select(driver.find_element_by_id('userinfo-year')).select_by_value('2018')
            time.sleep(2)

            # 设置学生类型
            Select(driver.find_element_by_id('userinfo-student_type')).select_by_value('studentType.STVS')
            time.sleep(2)
            # ((JavaScriptExecutor) driver).executeScript($("input#{放置元素的CLASS}[readonly]").attr("readonly", null);

            # 设置所属院校
            # js = 'document.getElementById("school_name").removeAttribute("readonly")'
            # driver.execute_script(js)
            # time.sleep(1)
            # driver.find_element_by_id("school_name").send_keys("济南职业学院")
            # time.sleep(2)

            # 设置所属院校
            driver.find_element_by_id('school_name').click()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="example_filter"]/label/input').send_keys(school)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="example"]/tbody/tr/td[3]/a').click()
            time.sleep(2)

            # 设置所属专业
            driver.find_element_by_id('userinfo-major_name').click()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="example2_filter"]/label/input').send_keys(professional)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="example2"]/tbody/tr/td[3]/a').click()
            time.sleep(2)

            # 设置所属专业
            # js = 'document.getElementById("userinfo-major_name").removeAttribute("readonly")'
            # driver.execute_script(js)
            # time.sleep(1)
            # driver.find_element_by_id("userinfo-major_name").send_keys("智能终端技术与应用")
            # time.sleep(2)

            # driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[2]/a').click()

            # 设置学号
            driver.find_element_by_id("userinfo-student_no").send_keys(random.randint(10000, 99999))
            time.sleep(2)

            # 设置详细地址
            driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
            time.sleep(2)

            driver.find_element_by_xpath('//*[@id="form-1"]/div[20]/button').click()
            time.sleep(3)

        print("账号: %s 完善个人信息结束*************************" % ac[0])




    """
    浏览资源
    """
    def BrowseTheResources(self,driver):

        # 进入资源中心，获取资源ID(按路径查找)
        # a='http://jinanzyk.36ve.com/ResourceCenter/resource/project-resource-list?projectId=890&page='+str(random.randint(1,500))+'&per-page=15&_pjax=%23new_html&_pjax=%23new_html'
        # driver.get(a)
        # time.sleep(3)
        # urls = driver.find_elements_by_xpath("//a")
        # time.sleep(2)
        # useSourceList=[]
        # for url in urls:
        #     if not (url.get_attribute("date-resource-id") is None):
        #         useSourceList.append(url.get_attribute("date-resource-id"))
        # time.sleep(2)

        # 进入资源中心，获取资源ID(按点击事件查找)
        driver.find_element_by_xpath('//*[@id="bs-example-navbar-collapse-1"]/ul/li[4]/a').click()
        time.sleep(1)

        while True:

            # 利用资源ID打开资源(按路径查找)（方法一）
            # course='http://jinanzyk.36ve.com/ResourceCenter/resource/show-resource?resource_id='+random.choice(useSourceList)
            # course='http://jinanzyk.36ve.com/ResourceCenter/resource/show-resource?resource_id=69c89e5e-7934-3360-bbb6-f688fc8485b7'
            # js = 'window.open("{}");'.format(course)
            # driver.execute_script(js)

            # 进入资源中心，获取资源ID(按点击事件查找)（方法二）
            # 挑选页数
            driver.find_element_by_xpath('//*[@id="new_html"]/div/div[3]/div[1]/div[3]/ul/div/input').clear()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="new_html"]/div/div[3]/div[1]/div[3]/ul/div/input').send_keys(str(random.randint(1, 600)))
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="courseUrl"]/span').click()
            time.sleep(2)

            resourceList = [i for i in range(1, 16)]
            for i in range(random.randint(5, 10)):

                # 挑选资源
                begin = '//*[@id="new_html"]/div/div[3]/div[1]/div[2]/div['
                end = ']/div/div/div[2]/h4/a'
                # 随机获取资源
                number = random.choice(resourceList)
                use = begin + str(number) + end
                # 移除资源，防止短时间内两次点击
                resourceList.remove(number)

                driver.find_element_by_xpath(use).click()
                time.sleep(2)

                # 切换窗口
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(1)

                # 获取资源类型（视频、PPT、word文档、JPG）
                dataType = driver.find_element_by_xpath('/html/body/div[1]/div/div[1]/h3').text
                print(dataType)
                time.sleep(3)

                # 视频类型
                if '.mp4' in dataType or '.avi' in dataType:
                    # 获取视频时长
                    needtime = driver.find_element_by_xpath('//*[@id="video"]/div/div[2]/div[8]').text
                    need = needtime.split('/')[1].strip().split(':')
                    needsecond = int(need[0]) * 60 + int(need[1])
                    time.sleep(1)

                    driver.find_element_by_xpath('//*[@id="video"]/div/div[9]/canvas').click()
                    stoptime = needsecond + random.randint(5, 10)
                    print("视频停留时长为 : %s" % stoptime)
                    time.sleep(stoptime)

                # ppt 类型
                elif '.pptx' in dataType or '.ppt' in dataType or '.pdf' in dataType:
                    # 获取其中的iframe
                    myiframe = driver.find_elements_by_tag_name('iframe')[0]
                    driver.switch_to.frame(myiframe)

                    countexit = login.isElementPresent(self, driver, 'id', 'PageCount')
                    if countexit:
                        # 获取ppt页数
                        pagecount = driver.find_element_by_id('PageCount').text

                        print("PPT 页数为 : %s" % pagecount)
                        driver.switch_to.default_content()

                        for i in range(int(pagecount) * 4):
                            myiframe.click()
                            time.sleep(random.randint(2, 4))

                # 文档类型
                elif '.docx' in dataType or '.doc' in dataType:

                    myiframe = driver.find_elements_by_tag_name('iframe')[0]
                    driver.switch_to.frame(myiframe)
                    result = login.isElementPresent(self, driver, 'id', 'ctn')
                    if result:

                        haha = login.isElementPresent(self, driver, 'id', 'pageCount')

                        if haha:
                            # 获取总页数
                            needCount = int(driver.find_element_by_id('pageCount').text)
                            print("获取数量: %s" % needCount)

                            # jsPrint='var allPage=document.querySelectorAll("#ctn>div");return allPage;'
                            # bodyHeight=driver.execute_script(jsPrint)
                            # time.sleep(1)
                            # print("窗口高度3: %s" % scroll)

                            for i in range(0, needCount * 1182, 150):
                                # 下滑操作
                                jsDown = "var q=document.getElementById('ctn').scrollTop=" + str(i)
                                driver.execute_script(jsDown)
                                time.sleep(1)
                # 其他类型
                else:
                    time.sleep(random.randint(60, 120))

                # 关闭窗口
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(2)



    """
    用来判断元素标签是否存在，
    """
    def isElementPresent(self,driver, by, value):
        try:
            element = driver.find_element(by=by, value=value)

        except NoSuchElementException as e:
            # 发生了NoSuchElementException异常，说明页面中未找到该元素，返回False
            return False
        else:
            # 没有发生异常，表示在页面中找到了该元素，返回True
            return True



    def chooseAddress(self):
        address=["山东","河南","河北","上海","北京","江苏","浙江","福建","广东","广西","四川","云南","贵州","重启","甘肃","山西","陕西"]
        return random.choice(address)


    def myreadExcel(self,filePath,school,professional):
        print(filePath)
        wb = openpyxl.load_workbook(filePath)
        # 获取所有工作表名
        names = wb.sheetnames
        # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
        sheet = wb[names[0]]
        # 获取最大行数
        maxRow = sheet.max_row
        # 获取最大列数
        maxColumn = sheet.max_column

        print("总行数：" + str(maxRow))
        print("总列数：" + str(maxColumn))

        countList = []

        for one_column_data in sheet.iter_rows():

            account=one_column_data[0].value

            password = 'znzd@123456'
            # password = 'znzd@12'
            need = (account, password)
            countList.append(need)

        # list2 = login.list_split(countList, 5)

        count = 1

        # 设置线程池
        # mypool=ThreadPoolExecutor(max_workers=5)

        # for i in list2:
        #
        #     for ac in i:
        #
        #         # print("编号 : %s , 账号 : %s" % (count, ac[0]))
        #
        #         mthread = threading.Thread(target=login.loginWebOne, args=(login, ac, school, professional))
        #         # mypool.submit(login.loginWebOne,ac, school, professional)
        #         # 启动刚刚创建的线程
        #         mthread.start()
        #
        #         count = count + 1
        #
        #     # time.sleep(random.randint(550, 600))



        for i in countList:


                # print("编号 : %s , 账号 : %s" % (count, ac[0]))

            mthread = threading.Thread(target=login.loginWebOne, args=(login, i, school, professional,count))
            # mypool.submit(login.loginWebOne,ac, school, professional)
            # 启动刚刚创建的线程
            mthread.start()

            count = count + 1

            # time.sleep(random.randint(550, 600))



    # 读数据
    # def read_xlrd(excelFile,filePath,school,professional) -> object:
    #
    #     data = xlrd.open_workbook(filePath)
    #
    #     print("工作表为：" + str(data.sheet_names()))
    #     list = data.sheet_names()
    #
    #     table = data.sheet_by_name(list[0])
    #
    #     print("总行数：" + str(table.nrows))
    #     print("总列数：" + str(table.ncols))
    #     line = int(table.nrows)
    #
    #     countList = []
    #
    #     for l in range(1, line):
    #         account = str(table.cell(l, 0).value)
    #         # pwd = str(table.cell(l, 1).value)
    #         # print("第%s行第一列的值为: %s" % (l, account))
    #         # need=(account,pwd)
    #
    #         password = 'znzd@123456'
    #         # password = 'znzd@12'
    #         need=(account,password)
    #         countList.append(need)
    #
    #     list2 = login.list_split(countList, 5)
    #
    #     count = 1
    #
    #     for i in list2:
    #         for ac in i:
    #
    #             print("编号 : %s , 账号 : %s" % (count, ac[0]))
    #
    #             mthread = threading.Thread(target=login.loginWebOne, args=(login, ac,school,professional))
    #             # 启动刚刚创建的线程
    #             mthread.start()
    #             count = count + 1
    #
    #         time.sleep(random.randint(550,600))
    #
    #



    def writeExcel(self,account):

        self.lock.acquire()

        try:
            base_path = r"D:\\"
            use_time = datetime.datetime.now().strftime('%Y-%m-%d')

            a = ['lili-' + use_time, ".xlsx"]

            filename = "".join(a)
            print(filename)
            sChildPath = os.path.join(base_path, filename)
            print("文件路径为 : %s " % sChildPath)

            # 判断文件是否存在
            if not os.path.exists(sChildPath):
                # 创建一个workbook 设置编码
                workbook = openpyxl.Workbook()
                worksheet = workbook.get_sheet_by_name('Sheet')
                worksheet.cell(1, 1).value = '账号'
                worksheet.cell(1, 2).value = '时间'
                workbook.save(sChildPath)

            data = openpyxl.load_workbook(sChildPath)

            # 取第一张表
            sheetnames = data.get_sheet_names()
            table = data.get_sheet_by_name(sheetnames[0])
            # table = data.active
            # print(table.title)  # 输出表名
            nrows = table.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数

            table.cell(nrows + 1, 1).value = account
            table.cell(nrows + 1, 2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            data.save(sChildPath)

        except Exception as e:
            print("====================异常信息=============================")
            print(traceback.format_exc())

        finally:
            self.lock.release()






if __name__ == '__main__':


    login.loginWebOne(login,None,None,None,1)
    # mytext='00:00 / 04:07'
    # need=mytext.split('/')[1].strip().split(':')
    # print(need)
    # needsecond=int(need[0])*60+int(need[1])
    # print(700/100)





