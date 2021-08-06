import os
import random
import threading
import traceback
import datetime
import xlrd
from selenium import webdriver
import time
import encodings.idna

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
import openpyxl


class login:

    lock=threading.Lock()

    # 均分数组
    def list_split(self, n):
        return [self[i:i + n] for i in range(0, len(self), n)]

    # 登录页面
    def loginWebOne(self, ac,school,professional):

        # ac = ('15083059097','znzd@123456')


        print("账号 %s 登陆开始===================================================》" % ac[0])


        # 笔记本路径
        root_path = os.path.abspath(os.path.dirname(__file__)).split('lili')[0]
        chromedriver = root_path+'images\chromedriver.exe'

        # json = requests.get("http://127.0.0.1:5000/get/").json()['proxy']

        myoptions = Options()

        myoptions.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"')
        # 设置代理

        # myoptions.add_argument('--proxy-server=' + json)

        myoptions.add_argument("--incognito")

        myoptions.add_experimental_option("excludeSwitches", ['enable-automation', 'enable-logging'])

        # myoptions.add_argument("--proxy-server=http://27.191.60.197:3256")

        url = "http://jinanzyk.36ve.com/login/login"



        # 打开一个浏览器窗口
        driver = webdriver.Chrome(
            executable_path=chromedriver,
            options=myoptions)
        driver.implicitly_wait(10)
        driver.set_window_size(800, 800)

        try:
            # 发送请求
            driver.get(url)
            time.sleep(4)

            driver.find_element_by_id("loginform-username").send_keys(ac[0])
            time.sleep(2)

            driver.find_element_by_id("loginform-password").send_keys(ac[1])
            time.sleep(2)

            driver.find_element_by_xpath('//*[@id="login-form-1"]/div[3]/button').click()

            nowurl = driver.current_url

            if nowurl == url:
                print("账号 %s 登陆失败" % ac[0])
                result_login=login.isElementPresent(self,driver,'xpath','//*[@id="login-form-1"]/div[1]/div[2]/div[2]/p')

                if result_login:
                    login.writeExcel(self,ac[0])
                    time.sleep(10)
                    driver.delete_all_cookies()
                    driver.quit()

            else:
                print("账号 %s 登陆成功" % ac[0])
                time.sleep(2)

                edit=login.isElementPresent(self,driver,'xpath','/html/body/div[4]/div/div/div/div[2]/a')
                print(edit)

                if edit:
                    driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[2]/a').click()

                    time.sleep(3)
                    driver.find_element_by_id("edit_info").click()
                    time.sleep(2)

                    # 设置名字
                    name = driver.find_element_by_id("userinfo-pet_name").get_attribute('value')
                    print("获取的名字为：%s" % name)
                    if name == '' or name is None:
                        driver.find_element_by_id("userinfo-pet_name").send_keys(
                            random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
                    time.sleep(2)

                    # 设置生日
                    month = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"]
                    day =  ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"
                        , "11", "12", "13", "14", "15", "16", "17", "18", "19", "20"
                        , "21", "22", "23", "24", "25", "26", "27", "28", "29", "30"]
                    driver.find_element_by_id("userinfo-birthday").send_keys(
                        "2001-%s-%s" % (random.choice(month), random.choice(day)))
                    time.sleep(2)

                    name_flag = driver.find_element_by_xpath('//*[@id="userinfo-type"]').get_attribute('value')
                    print("%s 获取的身份为：%s" % (ac,name_flag))

                    if name_flag == "社会学习者":
                        driver.find_element_by_id("userinfo-company").send_keys(
                            random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
                        time.sleep(2)
                        driver.find_element_by_id("userinfo-pet_name").send_keys(
                            random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
                        time.sleep(2)
                        # 设置详细地址
                        driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
                        time.sleep(2)

                        driver.find_element_by_xpath('//*[@id="form-1"]/div[14]/button').click()

                    if name_flag == "企业用户":
                        driver.find_element_by_id("userinfo-company").send_keys(
                        random.sample('zyxwvutsrqponmlkjihgfedcba', 5))
                        time.sleep(2)
                        # 设置详细地址
                        driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
                        time.sleep(2)

                        driver.find_element_by_xpath('//*[@id="form-1"]/div[14]/button').click()

                    if name_flag == "学生":
                        # 设置入校年份
                        Select(driver.find_element_by_id('userinfo-year')).select_by_value('2018')
                        time.sleep(2)

                        # 设置学生类型
                        Select(driver.find_element_by_id('userinfo-student_type')).select_by_value('studentType.STVS')
                        time.sleep(2)
                        # ((JavaScriptExecutor) driver).executeScript($("input#{放置元素的CLASS}[readonly]").attr("readonly", null);

                        # 设置所属院校
                        # js = 'document.getElementById("school_name").removeAttribute("readonly")'
                        # driver.execute_script(js)
                        # time.sleep(1)
                        # driver.find_element_by_id("school_name").send_keys("济南职业学院")
                        # time.sleep(2)

                        # 设置所属院校
                        driver.find_element_by_id('school_name').click()
                        time.sleep(1)
                        driver.find_element_by_xpath('//*[@id="example_filter"]/label/input').send_keys(school)
                        time.sleep(1)
                        driver.find_element_by_xpath('//*[@id="example"]/tbody/tr/td[3]/a').click()
                        time.sleep(2)

                        # 设置所属专业
                        driver.find_element_by_id('userinfo-major_name').click()
                        time.sleep(1)
                        driver.find_element_by_xpath('//*[@id="example2_filter"]/label/input').send_keys(professional)
                        time.sleep(1)
                        driver.find_element_by_xpath('//*[@id="example2"]/tbody/tr/td[3]/a').click()
                        time.sleep(2)

                        # 设置所属专业
                        # js = 'document.getElementById("userinfo-major_name").removeAttribute("readonly")'
                        # driver.execute_script(js)
                        # time.sleep(1)
                        # driver.find_element_by_id("userinfo-major_name").send_keys("智能终端技术与应用")
                        # time.sleep(2)

                        # driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[2]/a').click()

                        # 设置学号
                        driver.find_element_by_id("userinfo-student_no").send_keys(random.randint(10000, 99999))
                        time.sleep(2)

                        # 设置详细地址
                        driver.find_element_by_id("userinfo-address").send_keys(login.chooseAddress(self))
                        time.sleep(2)

                        driver.find_element_by_xpath('//*[@id="form-1"]/div[20]/button').click()
                        time.sleep(3)


                    print("账号 %s 登陆结束*****************************************" % ac[0])

                time.sleep(3)

                a='http://jinanzyk.36ve.com/ResourceCenter/resource/project-resource-list?projectId=890&page='+str(random.randint(1,500))+'&per-page=15&_pjax=%23new_html&_pjax=%23new_html'
                driver.get(a)
                time.sleep(3)

                urls = driver.find_elements_by_xpath("//a")

                useSourceList=[]
                for url in urls:
                    if not (url.get_attribute("date-resource-id") is None):
                        useSourceList.append(url.get_attribute("date-resource-id"))
                    # print(url.get_attribute("date-resource-id"))

                # print(useSourceList)

                course='http://jinanzyk.36ve.com/ResourceCenter/resource/show-resource?resource_id='+random.choice(useSourceList)

                js = 'window.open("{}");'.format(course)

                driver.execute_script(js)
                time.sleep(5)

                # search_window = driver.current_window_handle
                driver.switch_to.window(driver.window_handles[-1])

                video=login.isElementPresent(self,driver,'xpath','//*[@id="video"]/div/div[9]/canvas')
                if video:
                    driver.find_element_by_xpath('//*[@id="video"]/div/div[9]/canvas').click()


        except Exception as e:
            print("======================%s 账号异常信息===============================" % ac[0])
            print(e)
            pass
        finally:
            time.sleep(random.randint(300,480))
            driver.delete_all_cookies()
            driver.quit()


    """
    用来判断元素标签是否存在，
    """
    def isElementPresent(self,driver, by, value):
        try:
            element = driver.find_element(by=by, value=value)
        # 原文是except NoSuchElementException, e:
        except NoSuchElementException as e:
            # 发生了NoSuchElementException异常，说明页面中未找到该元素，返回False
            return False
        else:
            # 没有发生异常，表示在页面中找到了该元素，返回True
            return True



    def chooseAddress(self):
        address=["山东","河南","河北","上海","北京","江苏","浙江","福建","广东","广西","四川","云南","贵州","重启","甘肃","山西","陕西"]
        return random.choice(address)



    # 读数据
    def read_xlrd(excelFile,filePath,school,professional) -> object:

        data = xlrd.open_workbook(filePath)

        print("工作表为：" + str(data.sheet_names()))
        list = data.sheet_names()

        table = data.sheet_by_name(list[0])

        print("总行数：" + str(table.nrows))
        print("总列数：" + str(table.ncols))
        line = int(table.nrows)

        countList = []

        for l in range(1, line):
            account = str(table.cell(l, 0).value)
            # pwd = str(table.cell(l, 1).value)
            # print("第%s行第一列的值为: %s" % (l, account))
            # need=(account,pwd)

            password = 'znzd@123456'
            # password = 'znzd@12'
            need=(account,password)
            countList.append(need)

        list2 = login.list_split(countList, 5)

        count = 1

        for i in list2:
            for ac in i:

                print("编号 : %s , 账号 : %s" % (count, ac[0]))

                mthread = threading.Thread(target=login.loginWebOne, args=(login, ac,school,professional))
                # 启动刚刚创建的线程
                mthread.start()
                count = count + 1

            time.sleep(random.randint(500,600))





    def writeExcel(self,account):

        self.lock.acquire()

        try:
            base_path = r"D:\\"
            use_time = datetime.datetime.now().strftime('%Y-%m-%d')

            a = ['lili-' + use_time, ".xlsx"]

            filename = "".join(a)
            print(filename)
            sChildPath = os.path.join(base_path, filename)
            print("文件路径为 : %s " % sChildPath)

            # 判断文件是否存在
            if not os.path.exists(sChildPath):
                # 创建一个workbook 设置编码
                workbook = openpyxl.Workbook()
                worksheet = workbook.get_sheet_by_name('Sheet')
                worksheet.cell(1, 1).value = '账号'
                worksheet.cell(1, 2).value = '时间'
                workbook.save(sChildPath)

            data = openpyxl.load_workbook(sChildPath)

            # 取第一张表
            sheetnames = data.get_sheet_names()
            table = data.get_sheet_by_name(sheetnames[0])
            # table = data.active
            # print(table.title)  # 输出表名
            nrows = table.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数

            table.cell(nrows + 1, 1).value = account
            table.cell(nrows + 1, 2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            data.save(sChildPath)

        except Exception as e:
            print("====================异常信息=============================")
            print(traceback.format_exc())

        finally:
            self.lock.release()






if __name__ == '__main__':

    # root_path = os.path.abspath(os.path.dirname(__file__)).split('lili')[0]
    # print(root_path)

    # list2 = login.read_xlrd(None)
    #
    # count = 1
    #
    # for i in list2:
    #     for ac in i:
    #         print("编号 : %s , 账号 : %s" % (count, ac))
    #
    #         mthread = threading.Thread(target=login.loginWebOne, args=(login, ac))
    #         # 启动刚刚创建的线程
    #         mthread.start()
    #         count = count + 1
    #
    #     time.sleep(60)
    #
    # print("账号全部结束，数量为 : %s" % count)

    # login.loginWebOne(login,None)



    # json=requests.get("http://127.0.0.1:5000/get/").json()['proxy']
    # print(json)

    # login.writeExcel(login,'4','3')
    login.loginWebOne(login,None,None,None)




