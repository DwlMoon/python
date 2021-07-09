import time
import re
import xlrd
import xlwt
import pymysql
import openpyxl


excel_read_file = r"E:\123.xlsx"
excel_account_path = r"data"

read_sheet_name = "sp"

table_title = ['u_cpf', 'u_email', 'u_passw', 'u_name', 'u_phone', 'u_street1', 'u_street2', 'u_uf_full', 'u_city', 'u_cep']

# [^**]	表示不匹配此字符集中的任何一个字符
# \u4e00-\u9fa5	汉字的unicode范围
# \u0030-\u0039	数字的unicode范围
# \u0041-\u005a	大写字母unicode范围
# \u0061-\u007a	小写字母unicode范围
# \uAC00-\uD7AF	韩文的unicode范围
# \u3040-\u31FF	日文的unicode范围
my_reg = u"([^\u0030-\u0039\u0041-\u005a\u0061-\u007a])"

start_row = 1
end_row = 999999

# start_row = 10
# end_row = 20

table_value_array = []


def excel_task():
    print("---start---excel_task---")

    work_book = xlrd.open_workbook(excel_read_file)
    sheet_read = work_book.sheet_by_name(read_sheet_name)

    account_name = time.strftime('%Y-%m-%d_%H-%M-%S')
    u_time = time.strftime('%Y/%m/%d %H:%M:%S')
    print(f"---u_time={u_time}")

    excel_account_file = rf"E:\account_{account_name}.xlsx"
    print(f"---excel_account_file={excel_account_file}")

    table_value_array.append(table_title)

    for row in range(start_row, end_row):
        table_line_value_array_new = []

        table_line_value_array = sheet_read.row_values(row)
        column_value_0 = table_line_value_array[0]
        u_name_lower = column_value_0.lower()
        u_name = re.sub(my_reg, "", u_name_lower)

        u_cpf = table_line_value_array[1]

        'u_cpf', 'u_email', 'u_passw', 'u_name', 'u_phone', 'u_street1', 'u_street2', 'u_uf_full', 'u_city', 'u_cep'

        table_line_value_array_new.append(u_cpf)
        table_line_value_array_new.append(u_email)
        table_line_value_array_new.append(u_passw)
        table_line_value_array_new.append(u_name)
        table_line_value_array_new.append(u_phone)
        table_line_value_array_new.append(u_street1)
        table_line_value_array_new.append(u_street2)
        table_line_value_array_new.append(u_uf_full)
        table_line_value_array_new.append(u_city)
        table_line_value_array_new.append(u_cep)

        table_value_array.append(table_line_value_array_new)
        pass
    print(f"---table_value---读取成功---{end_row-start_row}")

    print("---start---create---table---")
    f = openpyxl.Workbook()
    f_sheet = f.active

    for row in range(len(table_value_array)):
        f_sheet.append(table_value_array[row])
        pass

    f.save(excel_account_file)
    print("---save---table_data---")
    pass


def read_excel_2():
    # 打开文件
    work_book = xlrd.open_workbook(excel_read_file)

    # 获取所有sheet
    # sheet_name = work_book.sheet_names()[0]

    # 根据sheet索引或者名称获取sheet内容
    # sheet = work_book.sheet_by_index(0)

    sheet_read = work_book.sheet_by_name(read_sheet_name)

    # 获取指定单元格里面的值
    # sheet.cell_value(第几行, 第几列)

    # 获取整行和整列的值（数组）
    # 获取第2行内容
    # rows = sheet.row_values(1)
    # 获取第3列内容
    # cols = sheet.col_values(2)

    f = openpyxl.Workbook()
    f_sheet = f.active

    print("---start---add---table_title---")
    for col in range(len(table_title)):
        c = col + 1
        c_value = table_title[col]
        f_sheet.cell(row=1, column=c, value=c_value)
        pass
    print("---end---add---table_title---")
    # 获取sheet的名称，行数，列数
    # print(sheet.name, sheet.nrows, sheet.ncols)

    print("---start---add---table_data---")
    for row in range(start_row, end_row):
        rows_value_line = sheet_read.row_values(row)
        column_value_0 = rows_value_line[0]
        u_name_lower = column_value_0.lower()
        u_name = re.sub(my_reg, "", u_name_lower)

        column_value_1 = rows_value_line[1]



        # for col in range(len(table_title)):
        #     my_row =
        #     my_col = col + 1
        #     f_sheet.cell(row=2, column=my_col, value=sub_str)

    print("---end---add---table_data---")

    # f.save(excel_account_path)
    print("---save---table_data---")
    pass


def read_excel():
    # 打开文件
    work_book = xlrd.open_workbook(excel_read_file)

    # 1.获取sheet的名字
    # 1.1 获取所有sheet的名字(list类型)
    all_sheet_names = work_book.sheet_names()
    print(all_sheet_names)

    # 1.2 按索引号获取sheet的名字（string类型）
    sheet1Name = work_book.sheet_names()[0]
    print(sheet1Name)

    # 2. 获取sheet内容
    ## 2.1 法1：按索引号获取sheet内容
    # sheet索引从0开始
    sheet1_content1 = work_book.sheet_by_index(0)
    ## 2.2 法2：按sheet名字获取sheet内容
    sheet1_content2 = work_book.sheet_by_name('Sheet1')

    # 3. sheet的名称，行数，列数
    print(sheet1_content1.name, sheet1_content1.nrows, sheet1_content1.ncols)

    # 4. 获取整行和整列的值（数组）
    # 获取第四行内容
    rows = sheet1_content1.row_values(3)
    # 获取第三列内容
    cols = sheet1_content1.col_values(2)
    print(rows)

    # 5. 获取单元格内容(三种方式)
    print(sheet1_content1.cell(1, 0).value)
    print(sheet1_content1.cell_value(2, 2))
    print(sheet1_content1.row(2)[2].value)

    # 6. 获取单元格内容的数据类型
    # Tips: python读取excel中单元格的内容返回的有5种类型 [0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error]
    print(sheet1_content1.cell(1, 0).ctype)
    pass


# def my_sql():
#     conn = pymysql.connect(host='superchat.cy1brcevuvyy.sa-east-1.rds.amazonaws.com', user='superchatmg',
#                            password='p8I9$yf6xNnzvig6', database='influuchatbackstage', port=3306)
#     cursor = conn.cursor()
#     # 若id选择自动递增并为主键，可以设为null,让其自动增长。
#     sql = """
#         insert into t_shop_data(pk_data_id,u_name,u_cpf,u_state,update_time,create_time) value(null,%s,%s,%s)
#         """
#     u_name = 'ADRIANA SAMPAIO'.replace(" ", "").lower()
#
#     print(f"u_name={u_name}")
#     u_cpf = '10907926894'
#
#     u_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
#
#     print(f"u_time={u_time}")
#
#     # insert_sql = f"insert into t_shop_data(u_name,u_cpf,update_time) value(null,{u_name},{u_cpf},now())"
#     insert_sql = f"insert into t_shop_data(u_name,u_cpf,update_time,create_time) value(null,{u_name},{u_cpf},{pymysql.escape_string(u_time)},{pymysql.escape_string(u_time)}) "
#     cursor.execute(insert_sql)
#     conn.commit()
#
#     conn.close()
#
#     pass


if __name__ == '__main__':
    # read_excel_2()
    excel_task()
    pass